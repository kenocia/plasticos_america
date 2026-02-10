import base64
from odoo import api, fields, models
import io
import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import xlsxwriter


class ReportInvoiceDetails(models.TransientModel):
    _name = 'kc_fiscal_hn.wizard.invoice.details'
    _description = 'Reporte Factura Detalle'

    company_id = fields.Many2one('res.company', 'Company', required=True,
                                 default=lambda s: s.env.company.id, index=True)
    fecha_desde = fields.Date(string="Fecha Desde", default=datetime.date.today(), required=True)
    fecha_hasta = fields.Date(string="Fecha Hasta", default=datetime.date.today(), required=True)

    def print_report(self):
        invoice = self.get_invoice()
        filename = 'Facturas - Producto de ' + str(self.fecha_desde) + ' a ' + str(self.fecha_hasta) + '.xlsx'

        # Crear un archivo Excel con xlsxwriter
        temp_output = io.BytesIO()
        workbook = xlsxwriter.Workbook(temp_output, {'in_memory': True})
        sheet = workbook.add_worksheet('Detalle Facturas')
        sheet2 = workbook.add_worksheet('Detalle Productos')
        sheet3 = workbook.add_worksheet('Margen por Producto')

        # Definir estilos
        encabezado = workbook.add_format({'bold': True, 'font_color': 'black', 'font_size': 14, 'align': 'center'})
        encabezado2 = workbook.add_format(
            {'bold': True, 'font_color': 'black', 'font_size': 16, 'align': 'center', 'valign': 'vcenter'})
        subtitulos = workbook.add_format({'bold': True, 'font_color': 'black', 'align': 'center'})
        detalle = workbook.add_format({'font_color': 'black'})
        detalle_moneda = workbook.add_format({'num_format': 'L#,##0.00'})
        detalle_porcentaje = workbook.add_format({'num_format': '0.00%'})

        # ENCABEZADO
        sheet.merge_range('B10:H10', 'Facturas Detalle', encabezado)
        sheet.merge_range('D1:D3', 'GRUPO RIO S. DE R.L.', encabezado2)
        sheet.merge_range('D4:D6', str(self.fecha_desde) + ' a ' + str(self.fecha_hasta), encabezado2)
        sheet2.merge_range('B10:H10', 'Productos Detalle', encabezado)
        sheet2.merge_range('D1:D3', 'GRUPO RIO S. DE R.L.', encabezado2)
        sheet2.merge_range('D4:D6', str(self.fecha_desde) + ' a ' + str(self.fecha_hasta), encabezado2)
        sheet3.merge_range('B10:H10', 'Margen por Producto', encabezado)
        sheet3.merge_range('D1:D3', 'GRUPO RIO S. DE R.L.', encabezado2)
        sheet3.merge_range('D4:D6', str(self.fecha_desde) + ' a ' + str(self.fecha_hasta), encabezado2)

        # TITULOS
        titles = ['N° Correlativo', 'Fecha', 'Cliente', 'Vendedor', 'Subtotal', 'Total', 'Plazos de Pago', 'Importe exento',
                  'Importe exonerado', 'Importe base 15%', 'Importe base 18%', 'Isv 15%','Isv 18%','Estado']
        for col, title in enumerate(titles, 1):
            sheet.write(11, col, title, subtitulos)

        f = 12
        subtotal = 0
        total = 0
        for i in invoice:
            subtotal += i['subtotal']
            total += i['total']
            sheet.write(f, 1, i['correlativo'], detalle)
            sheet.write(f, 2, i['fecha'], detalle)
            sheet.write(f, 3, i['cliente'], detalle)
            sheet.write(f, 4, i['vendedor'], detalle)
            sheet.write(f, 5, i['subtotal'], detalle_moneda)
            sheet.write(f, 6, i['total'], detalle_moneda)
            sheet.write(f, 7, i['plazo_pago'], detalle)
            sheet.write(f, 8, i['importe_exento'], detalle_moneda)
            sheet.write(f, 9, i['importe_exonerado'], detalle_moneda)
            sheet.write(f, 10, i['importe_isv15'], detalle_moneda)
            sheet.write(f, 11, i['importe_isv18'], detalle_moneda)
            sheet.write(f, 12, i['isv15'], detalle_moneda)
            sheet.write(f, 13, i['isv18'], detalle_moneda)
            sheet.write(f, 14, i['estado'], detalle)
            f += 1
        sheet.write(f, 4, 'TOTAL', detalle_moneda)
        sheet.write(f, 5, subtotal, detalle_moneda)
        sheet.write(f, 6, total, detalle_moneda)
        # ---------------------------- HOJA 2 -----------------------------------------------

        details = self.get_details_invoice()

        titles = ['N° Correlativo', 'Fecha', 'Cliente', 'Vendedor', 'Cantidad', 'Producto', 'Code', 'Precio Unitario',
                  'Costo Unitario', 'Costo Total', 'Importe exento', 'Importe exonerado', 'Importe base 15%',
                  'Importe base 18%', 'Isv 15%', 'Isv 18%', 'Subtotal', 'Total']
        for col, title in enumerate(titles, 1):
            sheet2.write(11, col, title, subtitulos)

        subtotald = 0
        totald = 0
        fd = 12
        for i in details:
            sheet2.write(fd, 1, i['correlativo'], detalle)
            sheet2.write(fd, 2, i['fecha'], detalle)
            sheet2.write(fd, 3, i['cliente'], detalle)
            sheet2.write(fd, 4, i['vendedor'], detalle)
            sheet2.write(fd, 5, i['cantidad'], detalle)
            sheet2.write(fd, 6, i['producto'], detalle)
            sheet2.write(fd, 7, i['code'], detalle)
            sheet2.write(fd, 8, i['precio'], detalle_moneda)
            sheet2.write(fd, 9, i['costo'], detalle_moneda)
            sheet2.write(fd, 10, i['costo_total'], detalle_moneda)
            sheet2.write(fd, 11, i['importe_exento'], detalle_moneda)
            sheet2.write(fd, 12, i['importe_exonerado'], detalle_moneda)
            sheet2.write(fd, 13, i['importe_isv15'], detalle_moneda)
            sheet2.write(fd, 14, i['importe_isv18'], detalle_moneda)
            sheet2.write(fd, 15, i['isv15'], detalle_moneda)
            sheet2.write(fd, 16, i['isv18'], detalle_moneda)
            sheet2.write(fd, 17, i['subtotal'], detalle_moneda)
            sheet2.write(fd, 18, i['total'], detalle_moneda)
            subtotald += i['subtotal']
            totald += i['total']
            fd += 1
        sheet2.write(fd, 16, 'TOTAL', detalle_moneda)
        sheet2.write(fd, 17, subtotald, detalle_moneda)
        sheet2.write(fd, 18, totald, detalle_moneda)

        # ---------------------------- HOJA 3 -----------------------------------------------

        margen = self.get_margen_invoice()

        titles = ['Producto', 'Code', 'Importe exento', 'Importe exonerado', 'Importe base 15%', 'Importe base 18%', 'Isv 15%', 'Isv 18%',
                  'Subtotal', 'Total', 'Costo Total', 'Margen de Contribucion', '% Margen Total', '% Margen x Producto']
        for col, title in enumerate(titles, 1):
            sheet3.write(11, col, title, subtitulos)

        importe_exento = 0
        importe_exonerado = 0
        importe_isv15 = 0
        importe_isv18 = 0
        isv15 = 0
        isv18 = 0
        subtotal = 0
        total = 0
        costo_total = 0
        margen_contribucion = 0
        margen_total = 0
        margen_producto = 0
        line_margen_total = 0
        line_margen_producto = 0
        fd = 12
        for i in margen:
            try:
                line_margen_producto = i['margen_contribucion'] / i['subtotal']
            except ZeroDivisionError:
                line_margen_producto = 0
            sheet3.write(fd, 1, i['producto'], detalle)
            sheet3.write(fd, 2, i['code'], detalle)
            sheet3.write(fd, 3, i['importe_exento'], detalle_moneda)
            sheet3.write(fd, 4, i['importe_exonerado'], detalle_moneda)
            sheet3.write(fd, 5, i['importe_isv15'], detalle_moneda)
            sheet3.write(fd, 6, i['importe_isv18'], detalle_moneda)
            sheet3.write(fd, 7, i['isv15'], detalle_moneda)
            sheet3.write(fd, 8, i['isv18'], detalle_moneda)
            sheet3.write(fd, 9, i['subtotal'], detalle_moneda)
            sheet3.write(fd, 10, i['total'], detalle_moneda)
            sheet3.write(fd, 11, i['costo_total'], detalle_moneda)
            sheet3.write(fd, 12, i['margen_contribucion'], detalle_moneda)
            sheet3.write(fd, 13, i['margen_total'], detalle_porcentaje)
            sheet3.write(fd, 14, line_margen_producto, detalle_porcentaje)
            importe_exento += i['importe_exento']
            importe_exonerado += i['importe_exonerado']
            importe_isv15 += i['importe_isv15']
            importe_isv18 += i['importe_isv18']
            subtotal += i['subtotal']
            total += i['total']
            costo_total += i['costo_total']
            margen_contribucion += i['margen_contribucion']
            margen_total += i['margen_total']
            fd += 1
        sheet3.write(fd, 2, 'TOTAL', detalle)
        sheet3.write(fd, 3, importe_exento, detalle_moneda)
        sheet3.write(fd, 4, importe_exonerado, detalle_moneda)
        sheet3.write(fd, 5, importe_isv15, detalle_moneda)
        sheet3.write(fd, 6, importe_isv18, detalle_moneda)
        sheet3.write(fd, 7, isv15, detalle_moneda)
        sheet3.write(fd, 8, isv18, detalle_moneda)
        sheet3.write(fd, 9, subtotal, detalle_moneda)
        sheet3.write(fd, 10, total, detalle_moneda)
        sheet3.write(fd, 11, costo_total, detalle_moneda)
        sheet3.write(fd, 12, margen_contribucion, detalle_moneda)
        sheet3.write(fd, 13, margen_total, detalle_porcentaje)

        # Cerrar el archivo
        workbook.close()

        # Leer el archivo con openpyxl para insertar la imagen
        temp_output.seek(0)
        final_output = io.BytesIO()
        workbook = load_workbook(temp_output)

        # Obtener el logo de la compañía
        company = self.company_id
        if company.logo:
            logo_data = base64.b64decode(company.logo)
            for sheet_name in workbook.sheetnames:
                logo_image = io.BytesIO(logo_data)  # Crear una nueva instancia de BytesIO para cada hoja
                img = OpenpyxlImage(logo_image)
                sheet = workbook[sheet_name]
                sheet.add_image(img, 'A1')

        # Guardar el archivo final
        workbook.save(final_output)
        final_output.seek(0)

        # Crear registro para exportar el archivo
        export_id = self.env['kc_fiscal_hn.invoice.details.excel'].create(
            {'excel_file': base64.b64encode(final_output.getvalue()), 'file_name': filename})

        final_output.close()
        temp_output.close()

        return {
            'view_mode': 'form',
            'res_id': export_id.id,
            'res_model': 'kc_fiscal_hn.invoice.details.excel',
            'view_type': 'form',
            'type': 'ir.actions.act_window',
            'target': 'new'
        }

    def get_invoice(self):
        domain = [('company_id', '=', self.company_id.id),
                  ('invoice_date', '>=', self.fecha_desde),
                  ('invoice_date', '<=', self.fecha_hasta),
                  ('state', 'in', ['posted', 'cancel']),
                  ('move_type', 'in', ['out_invoice', 'out_refund'])]

        invoices = self.env['account.move'].search(domain)

        data = []
        for i in invoices:
            sign = 1 if i.move_type == 'out_invoice' else -1
            subtotal = sign * i.amount_untaxed if i.state == 'posted' else 0.00
            total = sign * i.amount_total if i.state == 'posted' else 0.00
            importe_exento = sign * i.amount_exento if i.state == 'posted' else 0.00
            importe_exonerado = sign * i.amount_exonerado if i.state == 'posted' else 0.00
            importe_isv15 = sign * i.gravado_isv15 if i.state == 'posted' else 0.00
            importe_isv18 = sign * i.gravado_isv18 if i.state == 'posted' else 0.00
            isv_15 = sign * i.amount_isv15 if i.state == 'posted' else 0.00
            isv_18 = sign * i.amount_isv18 if i.state == 'posted' else 0.00

            invoice = {
                'correlativo': i.name,
                'fecha': i.date.strftime('%d-%m-%Y') if i.date else '',
                'cliente': i.partner_id.name,
                'vendedor': i.invoice_user_id.name,
                'subtotal': subtotal,
                'total': total,
                'plazo_pago': i.invoice_payment_term_id.name,
                'importe_exento': importe_exento,
                'importe_exonerado': importe_exonerado,
                'importe_isv15': importe_isv15,
                'importe_isv18': importe_isv18,
                'isv15': isv_15,
                'isv18': isv_18,
                'estado': i.state
            }
            data.append(invoice)
        return data

    def get_details_invoice(self):
        domain = [('company_id', '=', self.company_id.id),
                  ('invoice_date', '>=', self.fecha_desde),
                  ('invoice_date', '<=', self.fecha_hasta),
                  ('product_id', '!=', False),
                  ('display_type', '=', 'product'),
                  ('parent_state', 'in', ['posted', 'cancel']),
                  ('move_type', 'in', ['out_invoice', 'out_refund'])]

        details = self.env['account.move.line'].search(domain)

        data = []
        for line in details:
            sign = 1 if line.move_id.move_type == 'out_invoice' else -1
            cantidad = sign * line.quantity if line.parent_state == 'posted' else 0.00
            precio = line.price_unit if line.parent_state == 'posted' else 0.00
            costo = line.product_id.standard_price if line.parent_state == 'posted' else 0.00
            costo_total = costo * cantidad

            # Calcular el subtotal con descuento
            subtotal = cantidad * precio
            descuento = (subtotal * line.discount) / 100 if line.discount else 0
            total = subtotal - descuento

            # Determinar si la línea es exenta o exonerada basado en la lógica de _compute_exent
            importe_exento = total if any(
                tax.tax_group_id.name == 'Exento' for tax in line.tax_ids) else 0.00
            importe_exonerado = total if any(
                tax.tax_group_id.name == 'Exonerado' for tax in line.tax_ids) else 0.00

            # Calcular ISV 15% si aplica (cuando no es exento ni exonerado)
            importe_isv15 = ((total) * 15 / 100) if any(
                tax.tax_group_id.name == '15%' for tax in
                line.tax_ids) and not importe_exento and not importe_exonerado else 0.00

            # Calcular ISV 18% si aplica (cuando no es exento ni exonerado)
            importe_isv18 = ((total) * 18 / 100) if any(
                tax.tax_group_id.name == '18%' for tax in
                line.tax_ids) and not importe_exento and not importe_exonerado else 0.00

            # Calcular ISV 15% si aplica (cuando no es exento ni exonerado)
            isv15 = ((total) * 15) if any(
                tax.tax_group_id.name == '15%' for tax in
                line.tax_ids) and not importe_exento and not importe_exonerado else 0.00

            # Calcular ISV 18% si aplica (cuando no es exento ni exonerado)
            isv18 = ((total) * 18) if any(
                tax.tax_group_id.name == '18%' for tax in
                line.tax_ids) and not importe_exento and not importe_exonerado else 0.00

            # Aplicar el signo a los impuestos
            importe_isv15 *= sign
            importe_isv18 *= sign

            # Sumar impuestos al total
            total += importe_isv15 + importe_isv18

            invoice_detail = {
                'correlativo': line.move_id.name,
                'fecha': line.date.strftime('%d-%m-%Y') if line.date else '',
                'cliente': line.partner_id.name,
                'vendedor': line.move_id.invoice_user_id.name,
                'cantidad': cantidad,
                'producto': line.product_id.name,
                'code': line.product_id.default_code,
                'precio': precio,
                'costo': costo,
                'costo_total': sign * costo_total,
                'importe_exento': sign * importe_exento,
                'importe_exonerado': sign * importe_exonerado,
                'importe_isv15': sign * importe_isv15,
                'importe_isv18': sign * importe_isv18,
                'isv15': sign * isv15,
                'isv18': sign * isv18,
                'subtotal': sign * subtotal,
                'total': sign * total
            }
            data.append(invoice_detail)

        return data

    def get_margen_invoice(self):
        domain = [('company_id', '=', self.company_id.id),
                  ('invoice_date', '>=', self.fecha_desde),
                  ('invoice_date', '<=', self.fecha_hasta),
                  ('product_id', '!=', False),
                  ('display_type', '=', 'product'),
                  ('parent_state', 'in', ['posted', 'cancel']),
                  ('move_type', 'in', ['out_invoice', 'out_refund'])]

        details = self.env['account.move.line'].search(domain)

        data = {}
        total_margen_contribucion = 0

        for line in details:
            sign = 1 if line.move_id.move_type == 'out_invoice' else -1
            cantidad = sign * line.quantity if line.parent_state == 'posted' else 0.00
            precio = sign * line.price_unit if line.parent_state == 'posted' else 0.00
            costo = line.product_id.standard_price if line.parent_state == 'posted' else 0.00
            costo_total = costo * cantidad

            # Inicializar total antes de utilizarlo
            subtotal = line.price_subtotal * sign if line.parent_state == 'posted' else 0.00
            total = subtotal  # Asegurar que total tenga un valor base

            # Inicializar valores de impuestos
            importe_exento = 0
            importe_exonerado = 0
            importe_isv15 = 0
            importe_isv18 = 0
            isv15 = 0
            isv18 = 0

            # Verificar si la línea es exenta o exonerada
            if any(tax.tax_group_id.name == 'Exento' for tax in line.tax_ids):
                importe_exento = total

            if any(tax.tax_group_id.name == 'Exonerado' for tax in line.tax_ids):
                importe_exonerado = total

            # Calcular ISV 15% si aplica
            if any(tax.tax_group_id.name == '15%' for tax in
                   line.tax_ids) and not importe_exento and not importe_exonerado:
                importe_isv15 = (total * 15 / 100)

            # Calcular ISV 18% si aplica
            if any(tax.tax_group_id.name == '18%' for tax in
                   line.tax_ids) and not importe_exento and not importe_exonerado:
                importe_isv18 = (total * 18 / 100)

            # Calcular ISV 15% si aplica
            if any(tax.tax_group_id.name == '15%' for tax in
                   line.tax_ids) and not importe_exento and not importe_exonerado:
                isv15 = (total * 15)

            # Calcular ISV 18% si aplica
            if any(tax.tax_group_id.name == '18%' for tax in
                   line.tax_ids) and not importe_exento and not importe_exonerado:
                isv18 = (total * 18)

            # Ajustar signos para impuestos
            importe_isv15 *= sign
            importe_isv18 *= sign

            # Calcular total sumando impuestos
            total += importe_isv15 + importe_isv18

            # Calcular margen de contribución
            margen_contribucion = subtotal - costo_total
            total_margen_contribucion += margen_contribucion

            # Agregar datos al diccionario por producto
            if line.product_id.id not in data:
                data[line.product_id.id] = {
                    'producto': line.product_id.name,
                    'code': line.product_id.default_code,
                    'importe_exento': 0,
                    'importe_exonerado': 0,
                    'importe_isv15': 0,
                    'importe_isv18': 0,
                    'isv15': 0,
                    'isv18': 0,
                    'subtotal': 0,
                    'total': 0,
                    'costo_total': 0,
                    'margen_contribucion': 0,
                    'margen_producto': 0,
                    'cantidad': 0
                }

            product_data = data[line.product_id.id]
            product_data['importe_exento'] += importe_exento
            product_data['importe_exonerado'] += importe_exonerado
            product_data['importe_isv15'] += importe_isv15
            product_data['importe_isv18'] += importe_isv18
            product_data['isv15'] += isv15
            product_data['isv18'] += isv18
            product_data['subtotal'] += subtotal
            product_data['total'] += total
            product_data['costo_total'] += costo_total
            product_data['margen_contribucion'] += margen_contribucion
            product_data['cantidad'] += cantidad

        # Calcular margen de contribución y margen total
        result = []
        for key, product_data in data.items():
            try:
                product_data['margen_producto'] = (product_data['margen_contribucion'] /
                                                   product_data['subtotal']) * 100
            except ZeroDivisionError:
                product_data['margen_producto'] = 0

            try:
                product_data['margen_total'] = product_data[
                                                   'margen_contribucion'] / total_margen_contribucion
            except ZeroDivisionError:
                product_data['margen_total'] = 0

            result.append(product_data)

        # Ordenar los resultados por margen de contribución de mayor a menor
        result_sorted = sorted(result, key=lambda x: x['margen_contribucion'],
                               reverse=True)

        return result_sorted


class ReportInvoiceDetailsExcel(models.TransientModel):
    _name = "kc_fiscal_hn.invoice.details.excel"
    _description = "Reporte Factura Detalle Excel"

    excel_file = fields.Binary('Lista Factura', readonly=True)
    file_name = fields.Char('Excel File', size=64)
